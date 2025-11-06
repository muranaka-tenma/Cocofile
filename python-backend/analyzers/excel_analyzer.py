"""
Excel分析モジュール

openpyxlを使用してExcelファイルからテキストを抽出します。
"""

import os
from typing import Dict, Any

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def analyze_excel(file_path: str) -> Dict[str, Any]:
    """
    Excelファイルを分析してテキストを抽出

    Args:
        file_path: Excelファイルのパス

    Returns:
        {
            "text": "抽出されたテキスト",
            "sheet_count": シート数,
            "file_size": ファイルサイズ(bytes)
        }
    """
    if not OPENPYXL_AVAILABLE:
        return {
            "error": "openpyxl is not installed",
            "text": "",
            "sheet_count": 0,
            "file_size": 0
        }

    if not os.path.exists(file_path):
        return {
            "error": f"File not found: {file_path}",
            "text": "",
            "sheet_count": 0,
            "file_size": 0
        }

    try:
        file_size = os.path.getsize(file_path)
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet_count = len(workbook.sheetnames)

        extracted_text = []

        # 全シートからテキスト抽出
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            extracted_text.append(f"[Sheet: {sheet_name}]")

            for row in sheet.iter_rows(values_only=True):
                row_text = " ".join([str(cell) for cell in row if cell is not None])
                if row_text.strip():
                    extracted_text.append(row_text)

        full_text = "\n".join(extracted_text)

        return {
            "text": full_text,
            "sheet_count": sheet_count,
            "file_size": file_size,
            "success": True
        }

    except Exception as e:
        return {
            "error": f"Failed to analyze Excel: {str(e)}",
            "text": "",
            "sheet_count": 0,
            "file_size": os.path.getsize(file_path) if os.path.exists(file_path) else 0
        }


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        result = analyze_excel(sys.argv[1])
        print(result)
    else:
        print("Usage: python excel_analyzer.py <excel_file_path>")
